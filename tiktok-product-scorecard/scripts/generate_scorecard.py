"""
TikTok Shop Product Scorecard Generator
Generates a formatted Excel workbook with:
- Scored comparison tab (if products provided)
- Individual product detail tabs (if products provided)
- Blank scorecard template tab (always)
- Visualization chart (if products provided)

Usage:
  python generate_scorecard.py --input products.json --output scorecard.xlsx [--chart chart.png]

Input JSON format:
[
  {
    "name": "Product Name",
    "category": "Product Category",
    "scores": [10, 9, 8, 7, 10, 9, 8, 9, 10, 7],
    "notes": ["note1", "note2", ..., "note10"]
  }
]

If --input is omitted, generates a blank scorecard only.
"""

import argparse
import json
import sys
import os

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ============================================================
# CONSTANTS
# ============================================================

CATEGORY_NAMES = [
    'Hook Potential',
    'Problem Solving Strength',
    'Demonstrability on Video',
    'Unique Selling Proposition',
    'Creator Compatibility',
    'Content Angle Depth',
    'Price Point & Margin Structure',
    'Operational Simplicity',
    'Emotional & Social Proof Potential',
    'Market Timing & Saturation',
]

QUESTIONS = [
    'Can it grab attention in the first 3-5 seconds?',
    'Does it solve a clear, emotional pain point?',
    'Can a creator show it in one clean shot?',
    'Distinct mechanism, design, or angle?',
    'Will creators WANT to make content with it?',
    'Can you generate 10+ angles, hooks, and case studies?',
    'Impulse price + room for affiliate, ads, margin?',
    'Easy to fulfill, low returns, FBT-friendly?',
    'Will the comment section explode?',
    'Is the category still scalable with whitespace?',
]

BANDS = [
    ('80-100', 'Aggressively Pursue', 'Greenlight full creator + ad rollout'),
    ('65-79', 'Strong w/ Proper Positioning', 'Focused creative angle; watch CPMs/CVR'),
    ('50-64', 'Needs Better Hooks', 'Rework positioning/packaging/pricing first'),
    ('Under 50', 'Hard to Scale', 'Walk away or rework the product'),
]

# Styles
HEADER_FONT = Font(name='Arial', bold=True, size=14, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1B2838', end_color='1B2838', fill_type='solid')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=11, color='1B2838')
CATEGORY_FONT = Font(name='Arial', bold=True, size=11, color='E8872B')
NORMAL_FONT = Font(name='Arial', size=10)
SCORE_FONT = Font(name='Arial', bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')


def add_conditional_formatting(ws, cell_ref):
    """Add score band conditional formatting to a cell."""
    ws.conditional_formatting.add(cell_ref,
        CellIsRule(operator='greaterThanOrEqual', formula=['80'], fill=GREEN_FILL))
    ws.conditional_formatting.add(cell_ref,
        CellIsRule(operator='between', formula=['65', '79'], fill=YELLOW_FILL))
    ws.conditional_formatting.add(cell_ref,
        CellIsRule(operator='between', formula=['50', '64'], fill=ORANGE_FILL))
    ws.conditional_formatting.add(cell_ref,
        CellIsRule(operator='lessThan', formula=['50'], fill=RED_FILL))


def create_comparison_sheet(wb, products):
    """Create multi-product comparison tab."""
    ws = wb.active
    ws.title = "Scored Comparison"
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35

    # Title
    ws.merge_cells(f'A1:{get_column_letter(2 + len(products))}1')
    ws['A1'] = 'TikTok Shop Product Scorecard — Scored Examples'
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30
    for col in range(1, 3 + len(products)):
        ws.cell(row=1, column=col).fill = HEADER_FILL

    # Product headers
    ws['A3'] = '#'
    ws['A3'].font = SUBHEADER_FONT
    ws['A3'].alignment = Alignment(horizontal='center')
    ws['B3'] = 'Category'
    ws['B3'].font = SUBHEADER_FONT

    for i, product in enumerate(products):
        col_idx = 3 + i
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
        cell = ws.cell(row=3, column=col_idx)
        cell.value = product['name']
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Category rows
    for i, name in enumerate(CATEGORY_NAMES):
        r = 4 + i
        ws[f'A{r}'] = i + 1
        ws[f'A{r}'].font = NORMAL_FONT
        ws[f'A{r}'].alignment = Alignment(horizontal='center')
        ws[f'B{r}'] = name
        ws[f'B{r}'].font = Font(name='Arial', bold=True, size=10)
        for j, product in enumerate(products):
            col_idx = 3 + j
            cell = ws.cell(row=r, column=col_idx)
            cell.value = product['scores'][i]
            cell.font = SCORE_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

    # Total row
    total_r = 14
    ws[f'B{total_r}'] = 'TOTAL SCORE'
    ws[f'B{total_r}'].font = Font(name='Arial', bold=True, size=12)
    for j in range(len(products)):
        col_idx = 3 + j
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=total_r, column=col_idx)
        cell.value = f'=SUM({col_letter}4:{col_letter}13)'
        cell.font = Font(name='Arial', bold=True, size=13)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='medium'), bottom=Side(style='medium'))
        add_conditional_formatting(ws, f'{col_letter}{total_r}')

    # Verdict row
    verdict_r = 15
    ws[f'B{verdict_r}'] = 'VERDICT'
    ws[f'B{verdict_r}'].font = Font(name='Arial', bold=True, size=11)
    for j in range(len(products)):
        col_idx = 3 + j
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=verdict_r, column=col_idx)
        cell.value = f'=IF({col_letter}{total_r}>=80,"PURSUE",IF({col_letter}{total_r}>=65,"STRONG",IF({col_letter}{total_r}>=50,"REWORK","PASS")))'
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def create_product_detail_sheet(wb, product):
    """Create individual product detail tab."""
    sheet_name = product['name'][:28]
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 70

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Product Scorecard: {product['name']}"
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = HEADER_FILL

    ws['A2'] = f"Category: {product.get('category', 'N/A')}"
    ws['A2'].font = Font(name='Arial', italic=True, size=10)

    # Headers
    ws['A4'] = '#'
    ws['A4'].font = SUBHEADER_FONT
    ws['A4'].alignment = Alignment(horizontal='center')
    ws['B4'] = 'Factor'
    ws['B4'].font = SUBHEADER_FONT
    ws['C4'] = 'Score'
    ws['C4'].font = SUBHEADER_FONT
    ws['C4'].alignment = Alignment(horizontal='center')
    ws['D4'] = 'Rationale'
    ws['D4'].font = SUBHEADER_FONT
    for col in range(1, 5):
        ws.cell(row=4, column=col).border = Border(bottom=Side(style='medium'))

    for i in range(10):
        r = 5 + i
        ws[f'A{r}'] = i + 1
        ws[f'A{r}'].font = NORMAL_FONT
        ws[f'A{r}'].alignment = Alignment(horizontal='center')
        ws[f'B{r}'] = CATEGORY_NAMES[i]
        ws[f'B{r}'].font = Font(name='Arial', bold=True, size=10)
        ws[f'C{r}'] = product['scores'][i]
        ws[f'C{r}'].font = SCORE_FONT
        ws[f'C{r}'].alignment = Alignment(horizontal='center')
        ws[f'C{r}'].border = THIN_BORDER
        ws[f'D{r}'] = product['notes'][i] if i < len(product.get('notes', [])) else ''
        ws[f'D{r}'].font = NORMAL_FONT
        ws[f'D{r}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 22

    # Total
    ws['B16'] = 'TOTAL'
    ws['B16'].font = Font(name='Arial', bold=True, size=12)
    ws['C16'] = '=SUM(C5:C14)'
    ws['C16'].font = Font(name='Arial', bold=True, size=14)
    ws['C16'].alignment = Alignment(horizontal='center')
    ws['C16'].border = Border(left=Side(style='medium'), right=Side(style='medium'),
                              top=Side(style='medium'), bottom=Side(style='medium'))
    ws['D16'] = '/ 100'
    ws['D16'].font = Font(name='Arial', bold=True, size=12)


def create_blank_scorecard(wb):
    """Create blank template tab for scoring new products."""
    ws = wb.create_sheet("Blank Scorecard")
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 50

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = 'TikTok Shop Product Scorecard — BLANK TEMPLATE'
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    for col in range(1, 6):
        ws.cell(row=1, column=col).fill = HEADER_FILL

    # Product info fields
    ws['A3'] = 'Product Name:'
    ws['A3'].font = SUBHEADER_FONT
    ws['B3'].border = Border(bottom=Side(style='thin'))
    ws['A4'] = 'Brand:'
    ws['A4'].font = SUBHEADER_FONT
    ws['B4'].border = Border(bottom=Side(style='thin'))
    ws['D3'] = 'Date:'
    ws['D3'].font = SUBHEADER_FONT
    ws['E3'].border = Border(bottom=Side(style='thin'))
    ws['D4'] = 'Scored By:'
    ws['D4'].font = SUBHEADER_FONT
    ws['E4'].border = Border(bottom=Side(style='thin'))

    # Column headers
    row = 6
    ws[f'A{row}'] = '#'
    ws[f'A{row}'].font = SUBHEADER_FONT
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'B{row}'] = 'Category'
    ws[f'B{row}'].font = SUBHEADER_FONT
    ws[f'C{row}'] = 'Scoring Question'
    ws[f'C{row}'].font = SUBHEADER_FONT
    ws[f'D{row}'] = 'Score (1-10)'
    ws[f'D{row}'].font = SUBHEADER_FONT
    ws[f'D{row}'].alignment = Alignment(horizontal='center')
    ws[f'E{row}'] = 'Notes / Rationale'
    ws[f'E{row}'].font = SUBHEADER_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = Border(bottom=Side(style='medium'))

    # Section 1
    ws.merge_cells('A7:E7')
    ws['A7'] = 'CATEGORIES 1-5: THE HOOK & CORE APPEAL'
    ws['A7'].font = CATEGORY_FONT

    score_cells = []
    start = 8
    for i in range(10):
        r = start + i
        if i == 5:
            ws.merge_cells(f'A{r}:E{r}')
            ws[f'A{r}'] = 'CATEGORIES 6-10: SCALABILITY & OPERATIONS'
            ws[f'A{r}'].font = CATEGORY_FONT
            start += 1
            r = start + i
        ws[f'A{r}'] = i + 1
        ws[f'A{r}'].font = NORMAL_FONT
        ws[f'A{r}'].alignment = Alignment(horizontal='center')
        ws[f'B{r}'] = CATEGORY_NAMES[i]
        ws[f'B{r}'].font = Font(name='Arial', bold=True, size=10)
        ws[f'C{r}'] = QUESTIONS[i]
        ws[f'C{r}'].font = NORMAL_FONT
        ws[f'D{r}'].border = THIN_BORDER
        ws[f'D{r}'].alignment = Alignment(horizontal='center')
        ws[f'E{r}'].border = THIN_BORDER
        ws.row_dimensions[r].height = 22
        score_cells.append(f'D{r}')

    # Total
    total_r = start + 11
    ws.merge_cells(f'A{total_r}:C{total_r}')
    ws[f'A{total_r}'] = 'TOTAL SCORE'
    ws[f'A{total_r}'].font = Font(name='Arial', bold=True, size=14)
    ws[f'A{total_r}'].alignment = Alignment(horizontal='right')
    ws[f'D{total_r}'] = '=' + '+'.join(score_cells)
    ws[f'D{total_r}'].font = Font(name='Arial', bold=True, size=16)
    ws[f'D{total_r}'].alignment = Alignment(horizontal='center')
    ws[f'D{total_r}'].border = Border(left=Side(style='medium'), right=Side(style='medium'),
                                       top=Side(style='medium'), bottom=Side(style='medium'))
    ws[f'E{total_r}'] = '/ 100'
    ws[f'E{total_r}'].font = Font(name='Arial', bold=True, size=14)
    add_conditional_formatting(ws, f'D{total_r}')

    # Verdict
    verdict_r = total_r + 2
    ws.merge_cells(f'A{verdict_r}:C{verdict_r}')
    ws[f'A{verdict_r}'] = 'VERDICT'
    ws[f'A{verdict_r}'].font = Font(name='Arial', bold=True, size=12)
    ws[f'A{verdict_r}'].alignment = Alignment(horizontal='right')
    ws.merge_cells(f'D{verdict_r}:E{verdict_r}')
    ws[f'D{verdict_r}'] = f'=IF(D{total_r}>=80,"AGGRESSIVELY PURSUE",IF(D{total_r}>=65,"STRONG w/ PROPER POSITIONING",IF(D{total_r}>=50,"NEEDS BETTER HOOKS","HARD TO SCALE")))'
    ws[f'D{verdict_r}'].font = Font(name='Arial', bold=True, size=12)
    ws[f'D{verdict_r}'].alignment = Alignment(horizontal='center')
    ws[f'D{verdict_r}'].border = THIN_BORDER

    # Score bands reference
    ref_r = verdict_r + 2
    ws.merge_cells(f'A{ref_r}:E{ref_r}')
    ws[f'A{ref_r}'] = 'SCORE BANDS REFERENCE'
    ws[f'A{ref_r}'].font = CATEGORY_FONT
    for i, (sr, v, a) in enumerate(BANDS):
        r = ref_r + 1 + i
        ws[f'A{r}'] = sr
        ws[f'A{r}'].font = Font(name='Arial', bold=True, size=10)
        ws[f'B{r}'] = v
        ws[f'B{r}'].font = Font(name='Arial', bold=True, size=10)
        ws.merge_cells(f'C{r}:E{r}')
        ws[f'C{r}'] = a
        ws[f'C{r}'].font = NORMAL_FONT


def create_chart(products, output_path):
    """Generate visualization chart comparing product scores."""
    try:
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        import numpy as np
    except ImportError:
        print("WARNING: matplotlib not installed. Skipping chart. Run: pip install matplotlib numpy")
        return False

    fig, axes = plt.subplots(2, 1, figsize=(14, 12))

    # Chart 1: Horizontal bar chart
    ax1 = axes[0]
    sorted_products = sorted(products, key=lambda x: sum(x['scores']), reverse=True)
    names = [p['name'] for p in sorted_products]
    totals = [sum(p['scores']) for p in sorted_products]

    colors = []
    for t in totals:
        if t >= 80: colors.append('#2E7D32')
        elif t >= 65: colors.append('#F9A825')
        elif t >= 50: colors.append('#E65100')
        else: colors.append('#C62828')

    bars = ax1.barh(range(len(names)), totals, color=colors, edgecolor='#1B2838', linewidth=0.5)
    ax1.set_yticks(range(len(names)))
    ax1.set_yticklabels(names, fontsize=11, fontweight='bold')
    ax1.set_xlim(0, 100)
    ax1.set_xlabel('Total Score (out of 100)', fontsize=11)
    ax1.set_title('TikTok Shop Product Scorecard — Total Scores', fontsize=14, fontweight='bold', pad=15)
    ax1.axvline(x=80, color='#2E7D32', linestyle='--', alpha=0.5)
    ax1.axvline(x=65, color='#F9A825', linestyle='--', alpha=0.5)
    ax1.axvline(x=50, color='#E65100', linestyle='--', alpha=0.5)

    for i, (bar, score) in enumerate(zip(bars, totals)):
        ax1.text(score + 1, i, str(score), va='center', fontsize=11, fontweight='bold')

    legend_patches = [
        mpatches.Patch(color='#2E7D32', label='80-100: Aggressively Pursue'),
        mpatches.Patch(color='#F9A825', label='65-79: Strong w/ Positioning'),
        mpatches.Patch(color='#E65100', label='50-64: Needs Better Hooks'),
        mpatches.Patch(color='#C62828', label='Under 50: Hard to Scale'),
    ]
    ax1.legend(handles=legend_patches, loc='lower right', fontsize=9)
    ax1.invert_yaxis()

    # Chart 2: Category breakdown for top 4
    ax2 = axes[1]
    top4 = sorted_products[:min(4, len(sorted_products))]
    x = np.arange(len(CATEGORY_NAMES))
    width = 0.2
    colors_top4 = ['#1B2838', '#E8872B', '#4A90D9', '#6B8E23']

    for i, product in enumerate(top4):
        offset = (i - 1.5) * width
        ax2.bar(x + offset, product['scores'], width, label=product['name'],
                color=colors_top4[i % len(colors_top4)], alpha=0.85)

    ax2.set_xticks(x)
    short_names = [n.replace(' & ', '\n& ').replace(' on ', '\non ') for n in CATEGORY_NAMES]
    ax2.set_xticklabels(short_names, fontsize=8, ha='center')
    ax2.set_ylabel('Score (1-10)', fontsize=10)
    ax2.set_ylim(0, 11)
    ax2.set_title('Category Breakdown — Top Products', fontsize=13, fontweight='bold', pad=15)
    ax2.legend(fontsize=9, loc='upper right')
    ax2.axhline(y=8, color='gray', linestyle=':', alpha=0.4)

    plt.tight_layout(pad=3)
    plt.savefig(output_path, dpi=150, bbox_inches='tight', facecolor='white', edgecolor='none')
    plt.close()
    return True


def main():
    parser = argparse.ArgumentParser(description='Generate TikTok Shop Product Scorecard')
    parser.add_argument('--input', '-i', help='JSON file with scored products')
    parser.add_argument('--output', '-o', default='TikTok_Shop_Product_Scorecard.xlsx',
                        help='Output Excel file path')
    parser.add_argument('--chart', '-c', help='Output chart image path (PNG)')
    args = parser.parse_args()

    products = []
    if args.input:
        with open(args.input, 'r') as f:
            products = json.load(f)
        # Validate
        for p in products:
            if len(p.get('scores', [])) != 10:
                print(f"ERROR: Product '{p.get('name')}' must have exactly 10 scores.")
                sys.exit(1)

    wb = Workbook()

    if products:
        create_comparison_sheet(wb, products)
        for product in products:
            create_product_detail_sheet(wb, product)
    else:
        wb.remove(wb.active)

    create_blank_scorecard(wb)
    wb.save(args.output)
    print(f"✅ Scorecard saved: {args.output}")

    if products and args.chart:
        if create_chart(products, args.chart):
            print(f"✅ Chart saved: {args.chart}")

    if not products:
        print("ℹ️  No products provided. Generated blank scorecard only.")
        print("   To score products, create a JSON file and pass with --input")


if __name__ == '__main__':
    main()
